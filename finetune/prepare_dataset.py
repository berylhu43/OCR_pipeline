#!/usr/bin/env python3
"""
Dataset Preparation Script for DeepSeek-OCR Fine-tuning

Converts Excel + Image pairs into a training dataset.
Each Excel row has a `page_num` column giving the exact document page it
belongs to, so matching is precise rather than approximate.

Data layout:
    training_data/
    ├── images/
    │   ├── 94. January 2023/   ← "paged" style: "January 2023 page N.jpeg"
    │   └── 95. February 2023/  ← "coded" style: "SECTION_NNN.jpeg"
    └── excel/
        ├── A.xlsx, A1.xlsx, ...

Output:
    training_data/dataset_train.jsonl
    training_data/dataset_val.jsonl
"""

import os
import json
import re
import argparse
import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl --break-system-packages --quiet")
    from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Section configuration
# ---------------------------------------------------------------------------

# Maps Excel stem → list of section codes used in image filenames.
# Most files map to a single code.  L_front maps to two codes because its
# data spans the L1 (page 111) and L2 (page 112) image sections.
SECTION_CONFIG: Dict[str, List[str]] = {
    "A":               ["A"],
    "A1":              ["A1"],
    "A2":              ["A2"],
    "B1":              ["B1"],
    "B2":              ["B2"],
    "C_beer":          ["C"],
    "C_cig":           ["C"],
    "C_liq":           ["C"],
    "D_beer":          ["D"],
    "D_cig":           ["D"],
    "D_liq":           ["D"],
    "E_Ames_total":    ["E"],
    "E_entres":        ["E"],
    "E_sorties":       ["E"],
    "F_Ames_Equip":    ["F"],
    "G":               ["G"],
    "H":               ["H"],
    "I_autorisation": ["I_autorisation"],
    "I_production":    ["I_production"],
    "J":               ["J"],
    "K":               ["K"],
    "L_front":         ["L1", "L2"],
    "L_combat":        ["L3"],
    "M1":              ["M1"],
    "M2":              ["M2"],
    "M3":              ["M3"],
    "M4":              ["M4"],
    "M5":              ["M5"],
}

# Page ranges for "paged"-style folders (e.g. January 2023).
# Used to build the section → {page_num: image_path} index when image files
# are named "Month Year page N.jpeg" rather than "SECTION_NNN.jpeg".
SECTION_PAGE_RANGES: Dict[str, Tuple[int, int]] = {
    "A":  (2, 2),
    "A1": (3, 14),
    "A2": (15, 23),
    "B1": (24, 35),
    "B2": (36, 49),
    "C":  (50, 52),
    "D":  (53, 87),
    "E":  (88, 90),
    "F":  (91, 91),
    "G":  (92, 96),
    "H":  (97, 106),
    "I_autorisation": (107, 107),
    "I_production":    (108, 108),
    "J":  (109, 109),
    "K":  (110, 110),
    "L1": (111, 111),
    "L2": (112, 112),
    "L3": (113, 114),
    "M1": (115, 115),
    "M2": (116, 119),
    "M3": (120, 125),
    "M4": (126, 136),
    "M5": (137, 142),
}

# ---------------------------------------------------------------------------
# Month name normalisation
# ---------------------------------------------------------------------------

_MONTH_VARIANTS: Dict[str, List[str]] = {
    "January":   ["january", "janvier", "janvr"],
    "February":  ["february", "feburary", "fevrier", "fevier", "février"],
    "March":     ["march", "mars"],
    "April":     ["april", "avril"],
    "May":       ["may", "mai"],
    "June":      ["june", "juin"],
    "July":      ["july", "juillet"],
    "August":    ["august", "aout", "août"],
    "September": ["september", "septembre"],
    "October":   ["october", "octobre"],
    "November":  ["november", "novembre"],
    "December":  ["december", "decembre", "décembre"],
}

MONTH_CANONICAL: Dict[str, str] = {}
for _canon, _alts in _MONTH_VARIANTS.items():
    MONTH_CANONICAL[_canon.lower()] = _canon
    for _a in _alts:
        MONTH_CANONICAL[_a.lower()] = _canon


def get_month_variants(canonical: str) -> Set[str]:
    return {k for k, v in MONTH_CANONICAL.items() if v == canonical}


# ---------------------------------------------------------------------------
# Image-folder discovery
# ---------------------------------------------------------------------------

def extract_month_folder_info(folder_name: str) -> Optional[Tuple[str, int]]:
    """Parse "94. January 2023" → ("January", 2023)."""
    m = re.search(r'(\w+)\s+(\d{4})$', folder_name)
    if not m:
        return None
    canon = MONTH_CANONICAL.get(m.group(1).lower())
    if not canon:
        return None
    return canon, int(m.group(2))


def detect_naming_style(folder_path: Path) -> str:
    """'coded' → A1_002.jpeg  |  'paged' → January 2023 page 3.jpeg  |  'numbered' → 1.jpg"""
    for f in folder_path.iterdir():
        if f.suffix.lower() in (".jpg", ".jpeg", ".png"):
            stem = f.stem
            if re.match(r'^\d+$', stem):
                return "numbered"
            if re.match(r'^.+_\d{3}$', stem):
                return "coded"
            if re.search(r'\bpage\s+\d+', stem, re.IGNORECASE):
                return "paged"
    return "coded"


# Section index type: section_code → {page_number → image_path}
SectionIndex = Dict[str, Dict[int, Path]]


def build_coded_index(folder_path: Path) -> SectionIndex:
    """
    Scan a coded folder and build section → {page_num → path}.
    A1_002.jpeg is automatically split off into section 'A' (page 2)
    so that A.xlsx gets the Effectifs summary page and A1.xlsx gets
    the recruit pages starting from page 3.
    """
    pat = re.compile(r'^(.+)_(\d+)\.(jpg|jpeg|png)$', re.IGNORECASE)
    index: SectionIndex = {}
    for f in folder_path.iterdir():
        m = pat.match(f.name)
        if m:
            code, page = m.group(1), int(m.group(2))
            index.setdefault(code, {})[page] = f

    # Peel page 2 off the A1 section → becomes the standalone 'A' section
    if "A1" in index and 2 in index["A1"]:
        index.setdefault("A", {})[2] = index["A1"].pop(2)

    return index


def build_numbered_index(folder_path: Path) -> SectionIndex:
    """
    Build section → {page_num → path} for numbered folders where images are
    simply named N.jpg (e.g. 1.jpg, 2.jpg, ..., 142.jpg).
    Uses SECTION_PAGE_RANGES to assign pages to sections.
    """
    page_to_path: Dict[int, Path] = {}
    for f in folder_path.iterdir():
        if f.suffix.lower() in (".jpg", ".jpeg", ".png") and f.stem.isdigit():
            page_to_path[int(f.stem)] = f

    index: SectionIndex = {}
    for code, (start, end) in SECTION_PAGE_RANGES.items():
        for page in range(start, end + 1):
            if page in page_to_path:
                index.setdefault(code, {})[page] = page_to_path[page]
    return index


def build_paged_index(
    folder_path: Path, month_name: str, year: int
) -> SectionIndex:
    """
    Build section → {page_num → path} for paged folders using
    SECTION_PAGE_RANGES and filenames like "January 2023 page N.jpeg".
    """
    index: SectionIndex = {}
    for code, (start, end) in SECTION_PAGE_RANGES.items():
        for page in range(start, end + 1):
            p = folder_path / f"{month_name} {year} page {page}.jpeg"
            if p.exists():
                index.setdefault(code, {})[page] = p
    return index


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def detect_date_columns(headers: tuple) -> Tuple[Optional[int], Optional[int]]:
    """Return (year_col_idx, month_col_idx); None when absent."""
    year_names  = {"annee", "year", "Year", "Annee", "ANNEE"}
    month_names = {"mois", "month", "Month", "Mois", "MOIS"}
    year_col = month_col = None
    for i, h in enumerate(headers):
        if h in year_names:
            year_col = i
        if h in month_names:
            month_col = i
    return year_col, month_col


def get_page_num_col(headers: tuple) -> Optional[int]:
    """Return index of the page_num column, or None if absent."""
    for i, h in enumerate(headers):
        if h == "page_num":
            return i
    return None


def get_data_rows(
    worksheet,
    month_variants: Set[str],
    year: int,
) -> Tuple[tuple, List[tuple]]:
    """
    Return (header_row, data_rows) filtered to the requested month/year.
    Skips the header row, secondary-header rows (non-numeric year), and
    rows from other months/years.
    """
    all_rows = [
        r for r in worksheet.iter_rows(values_only=True)
        if any(c is not None for c in r)
    ]
    if not all_rows:
        return (), []

    headers = all_rows[0]
    year_col, month_col = detect_date_columns(headers)

    data_rows = []
    for row in all_rows[1:]:
        if year_col is not None:
            yval = row[year_col]
            if not isinstance(yval, (int, float)):
                continue
            if not (2000 <= int(yval) <= 2100):
                continue
            if int(yval) != year:
                continue

        if month_col is not None and month_variants:
            mval = str(row[month_col]).strip() if row[month_col] is not None else ""
            if mval.lower() not in month_variants:
                continue

        data_rows.append(row)

    return headers, data_rows


def parse_page_num(value) -> List[int]:
    """
    Parse a page_num cell value into a list of page numbers.
      3              → [3]
      '[111, 112]'   → [111, 112]   (L_front spans two pages)
    """
    if isinstance(value, (int, float)):
        return [int(value)]
    if isinstance(value, str):
        nums = re.findall(r'\d+', value)
        return [int(n) for n in nums]
    return []


# ---------------------------------------------------------------------------
# Table formatting
# ---------------------------------------------------------------------------

def cell_to_str(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, float):
        return str(int(cell)) if cell == int(cell) else str(cell)
    if isinstance(cell, (datetime.datetime, datetime.date)):
        return cell.strftime("%Y-%m-%d")
    return str(cell).strip()


def rows_to_html(headers: tuple, data_rows: List[tuple]) -> str:
    if not data_rows:
        return ""
    max_cols = max(len(headers), max(len(r) for r in data_rows))

    def pad(row):
        return list(row) + [None] * (max_cols - len(row))

    parts = ["<table><tr>"]
    for h in pad(headers):
        parts.append(f"<td>{cell_to_str(h)}</td>")
    parts.append("</tr>")
    for row in data_rows:
        parts.append("<tr>")
        for cell in pad(row):
            parts.append(f"<td>{cell_to_str(cell)}</td>")
        parts.append("</tr>")
    parts.append("</table>")
    return "".join(parts)


def rows_to_markdown(headers: tuple, data_rows: List[tuple]) -> str:
    if not data_rows:
        return ""
    max_cols = max(len(headers), max(len(r) for r in data_rows))

    def pad(row):
        return list(row) + [None] * (max_cols - len(row))

    lines = [
        "| " + " | ".join(cell_to_str(c) for c in pad(headers)) + " |",
        "| " + " | ".join(["---"] * max_cols) + " |",
    ]
    for row in data_rows:
        lines.append("| " + " | ".join(cell_to_str(c) for c in pad(row)) + " |")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Training example creation
# ---------------------------------------------------------------------------

def create_training_example(image_path: Path, ground_truth: str, section_name: str = "") -> dict:
    if section_name:
        prompt = f"<image>\n<|grounding|>This is section {section_name}. The document is in Congolese French. Convert the document to markdown."
    else:
        prompt = "<image>\n<|grounding|>The document is in Congolese French. Convert the document to markdown."
    return {
        "image": str(image_path),
        "conversations": [
            {
                "role": "user",
                "content": prompt,
            },
            {
                "role": "assistant",
                "content": ground_truth,
            },
        ],
    }


# ---------------------------------------------------------------------------
# Per-month processing
# ---------------------------------------------------------------------------

def process_month_folder(
    folder_path: Path,
    excel_dir: Path,
    output_format: str = "html",
) -> List[dict]:
    """Process one month folder and return all training examples."""
    info = extract_month_folder_info(folder_path.name)
    if info is None:
        print(f"  Skipping '{folder_path.name}': cannot parse month/year")
        return []

    canonical_month, year = info
    naming_style = detect_naming_style(folder_path)
    month_variants = get_month_variants(canonical_month)

    print(f"  {canonical_month} {year}  |  style={naming_style}")

    # Build section → {page_num → image_path} index
    if naming_style == "coded":
        section_index = build_coded_index(folder_path)
    elif naming_style == "paged":
        section_index = build_paged_index(folder_path, canonical_month, year)
    else:  # numbered
        section_index = build_numbered_index(folder_path)

    print(f"  Sections found: {sorted(section_index)}")

    examples: List[dict] = []

    for excel_path in sorted(excel_dir.glob("*.xlsx")):
        stem = excel_path.stem
        if stem not in SECTION_CONFIG:
            continue

        # Merge page→image maps for all section codes this file covers
        page_to_img: Dict[int, Path] = {}
        for code in SECTION_CONFIG[stem]:
            page_to_img.update(section_index.get(code, {}))

        if not page_to_img:
            print(f"    {stem}: no images found – skipping")
            continue

        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb["Sheet1"]
        except Exception as e:
            print(f"    {stem}: load error – {e}")
            continue

        headers, data_rows = get_data_rows(ws, month_variants, year)
        if not data_rows:
            print(f"    {stem}: 0 rows for {canonical_month} {year} – skipping")
            continue

        pg_col = get_page_num_col(headers)
        if pg_col is None:
            print(f"    {stem}: no page_num column – skipping")
            continue

        # Strip page_num column for display; keep originals for grouping
        keep_cols = [i for i in range(len(headers)) if i != pg_col]
        display_headers = tuple(headers[i] for i in keep_cols)

        def strip_pg(row):
            return tuple(row[i] if i < len(row) else None for i in keep_cols)

        # Group original rows by their page_num value
        grouped: Dict[str, List[tuple]] = defaultdict(list)
        for row in data_rows:
            grouped[str(row[pg_col])].append(row)

        file_examples = 0
        for page_val_str, rows in grouped.items():
            pages = parse_page_num(page_val_str)
            imgs = [page_to_img[p] for p in pages if p in page_to_img]
            if not imgs:
                continue

            if len(imgs) == 1:
                # All rows for this page → one training example
                display_rows = [strip_pg(r) for r in rows]
                table = (
                    rows_to_html(display_headers, display_rows)
                    if output_format == "html"
                    else rows_to_markdown(display_headers, display_rows)
                )
                if table:
                    examples.append(create_training_example(imgs[0], table, stem))
                    file_examples += 1
            else:
                # Multi-page value (e.g. L_front '[111, 112]'):
                # distribute rows evenly across the images
                n_rows, n_imgs = len(rows), len(imgs)
                base, extra = divmod(n_rows, n_imgs)
                offset = 0
                for i, img in enumerate(imgs):
                    chunk = base + (1 if i < extra else 0)
                    chunk_rows = [strip_pg(r) for r in rows[offset: offset + chunk]]
                    offset += chunk
                    table = (
                        rows_to_html(display_headers, chunk_rows)
                        if output_format == "html"
                        else rows_to_markdown(display_headers, chunk_rows)
                    )
                    if table:
                        examples.append(create_training_example(img, table, stem))
                        file_examples += 1

        print(f"    {stem}: {len(data_rows)} rows → {file_examples} examples")

    return examples


# ---------------------------------------------------------------------------
# Dataset creation
# ---------------------------------------------------------------------------

def create_dataset(
    data_dir: Path,
    output_path: Path,
    output_format: str = "html",
    train_split: float = 0.9,
) -> Tuple[int, int]:
    images_dir = data_dir / "images"
    excel_dir  = data_dir / "excel"

    if not images_dir.exists():
        raise ValueError(f"Images directory not found: {images_dir}")
    if not excel_dir.exists():
        raise ValueError(f"Excel directory not found: {excel_dir}")

    month_folders = sorted(p for p in images_dir.iterdir() if p.is_dir())
    print(f"Found {len(month_folders)} month folder(s): {[f.name for f in month_folders]}")

    all_examples: List[dict] = []
    for folder in month_folders:
        print(f"\nProcessing: {folder.name}")
        examples = process_month_folder(folder, excel_dir, output_format)
        all_examples.extend(examples)
        print(f"  → {len(examples)} examples")

    print(f"\nTotal: {len(all_examples)} examples")
    if not all_examples:
        print("No examples created – check page_num columns and folder names.")
        return 0, 0

    import random
    random.shuffle(all_examples)
    split = int(len(all_examples) * train_split)
    train_ex, val_ex = all_examples[:split], all_examples[split:]

    train_path = output_path.parent / f"{output_path.stem}_train.jsonl"
    val_path   = output_path.parent / f"{output_path.stem}_val.jsonl"

    for path, subset in [(train_path, train_ex), (val_path, val_ex)]:
        with open(path, "w", encoding="utf-8") as f:
            for ex in subset:
                f.write(json.dumps(ex, ensure_ascii=False) + "\n")

    print(f"Saved {len(train_ex)} train  → {train_path}")
    print(f"Saved {len(val_ex)} val    → {val_path}")
    return len(train_ex), len(val_ex)


# ---------------------------------------------------------------------------
# Legacy helpers (kept for --test_single mode)
# ---------------------------------------------------------------------------

def excel_to_html_table(worksheet) -> str:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return ""
    non_empty = [r for r in rows if any(c is not None for c in r)]
    if not non_empty:
        return ""
    max_cols = max(len(r) for r in non_empty)

    def pad(row):
        return list(row) + [None] * (max_cols - len(row))

    parts = ["<table>"]
    for row in non_empty:
        parts.append("<tr>")
        for cell in pad(row):
            parts.append(f"<td>{cell_to_str(cell)}</td>")
        parts.append("</tr>")
    parts.append("</table>")
    return "".join(parts)


def create_single_example_from_existing(
    image_path: str,
    excel_path: str,
    sheet_name: str,
    output_path: str,
    output_format: str = "html",
):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    table = excel_to_html_table(ws)
    example = create_training_example(Path(image_path), table)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(example, f, indent=2, ensure_ascii=False)
    print(f"Created: {output_path}")
    print(f"\nExpected output (first 500 chars):\n{table[:500]}...")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Prepare dataset for DeepSeek-OCR fine-tuning"
    )
    parser.add_argument("--data_dir",    default="./training_data")
    parser.add_argument("--output",      default="./training_data/dataset.jsonl")
    parser.add_argument("--format",      choices=["html", "markdown"], default="html")
    parser.add_argument("--train_split", type=float, default=0.9)

    parser.add_argument("--test_single", action="store_true")
    parser.add_argument("--test_image")
    parser.add_argument("--test_excel")
    parser.add_argument("--test_sheet",  default="Sheet1")

    args = parser.parse_args()

    if args.test_single:
        if not args.test_image or not args.test_excel:
            print("For --test_single, provide --test_image and --test_excel")
        else:
            create_single_example_from_existing(
                args.test_image, args.test_excel, args.test_sheet,
                "test_example.json", args.format,
            )
    else:
        data_dir    = Path(args.data_dir)
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        create_dataset(data_dir, output_path, args.format, args.train_split)
